"""RPAForge Excel Library - Excel file operations."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING, Any

from rpaforge.core.activity import activity, library, output, tags

if TYPE_CHECKING:
    from pathlib import Path

logger = logging.getLogger("rpaforge.excel")


@library(name="Excel", category="Data", icon="📊")
class Excel:
    """Excel file operations library."""

    def __init__(self):
        self._workbook: Any = None
        self._workbook_path: str | None = None

    @property
    def _openpyxl(self):
        try:
            from openpyxl import Workbook, load_workbook

            return Workbook, load_workbook
        except ImportError as err:
            raise ImportError(
                "openpyxl is required for Excel library. "
                "Install it with: pip install rpaforge-libraries[excel]"
            ) from err

    @activity(name="Open Workbook", category="Excel")
    @tags("file", "open")
    @output("Path to the opened workbook")
    def open_workbook(
        self,
        path: str | Path,
        read_only: bool = False,
    ) -> str:
        _, load_workbook = self._openpyxl
        self._workbook = load_workbook(path, read_only=read_only)
        self._workbook_path = str(path)
        logger.info(f"Opened workbook: {path}")
        return self._workbook_path

    @activity(name="Create Workbook", category="Excel")
    @tags("file", "create")
    @output("Identifier for the new workbook")
    def create_workbook(self) -> str:
        Workbook, _ = self._openpyxl
        self._workbook = Workbook()
        self._workbook_path = None
        logger.info("Created new workbook")
        return "new_workbook"

    @activity(name="Close Workbook", category="Excel")
    @tags("file", "close")
    def close_workbook(self, save: bool = False) -> None:
        if self._workbook:
            try:
                if save and self._workbook_path:
                    self._workbook.save(self._workbook_path)
            finally:
                self._workbook.close()
                logger.info("Closed workbook")
        self._workbook = None
        self._workbook_path = None

    @activity(name="Save Workbook", category="Excel")
    @tags("file", "save")
    @output("Path where workbook was saved")
    def save_workbook(self, path: str | Path | None = None) -> str:
        if not self._workbook:
            raise ValueError("No workbook open")
        save_path = str(path) if path else self._workbook_path
        if not save_path:
            raise ValueError("No path specified for saving")
        self._workbook.save(save_path)
        self._workbook_path = save_path
        logger.info(f"Saved workbook: {save_path}")
        return save_path

    @activity(name="Get Sheet Names", category="Excel")
    @tags("sheet", "info")
    @output("List of sheet names")
    def get_sheet_names(self) -> list[str]:
        if not self._workbook:
            raise ValueError("No workbook open")
        return list(self._workbook.sheetnames)

    @activity(name="Get Active Sheet", category="Excel")
    @tags("sheet", "select")
    @output("Name of the active sheet")
    def get_active_sheet(self) -> str:
        if not self._workbook:
            raise ValueError("No workbook open")
        return self._workbook.active.title

    @activity(name="Set Active Sheet", category="Excel")
    @tags("sheet", "select")
    def set_active_sheet(self, name: str) -> None:
        if not self._workbook:
            raise ValueError("No workbook open")
        if name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' not found")
        self._workbook.active = self._workbook[name]
        logger.info(f"Set active sheet: {name}")

    @activity(name="Create Sheet", category="Excel")
    @tags("sheet", "create")
    @output("Name of the created sheet")
    def create_sheet(self, name: str, index: int | None = None) -> str:
        if not self._workbook:
            raise ValueError("No workbook open")
        if index is not None:
            sheet = self._workbook.create_sheet(name, index)
        else:
            sheet = self._workbook.create_sheet(name)
        logger.info(f"Created sheet: {name}")
        return sheet.title

    @activity(name="Delete Sheet", category="Excel")
    @tags("sheet", "delete")
    def delete_sheet(self, name: str) -> None:
        if not self._workbook:
            raise ValueError("No workbook open")
        if name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' not found")
        del self._workbook[name]
        logger.info(f"Deleted sheet: {name}")

    @activity(name="Read Cell", category="Excel")
    @tags("cell", "read")
    @output("Cell value")
    def read_cell(
        self,
        cell: str,
        sheet: str | None = None,
    ) -> Any:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        value = ws[cell].value
        logger.info(f"Read cell {cell}: {value}")
        return value

    @activity(name="Write Cell", category="Excel")
    @tags("cell", "write")
    def write_cell(
        self,
        cell: str,
        value: Any,
        sheet: str | None = None,
    ) -> None:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        ws[cell] = value
        logger.info(f"Wrote cell {cell}: {value}")

    @activity(name="Read Range", category="Excel")
    @tags("range", "read")
    @output("List of rows (or list of dicts if as_dict=True)")
    def read_range(
        self,
        range_spec: str,
        sheet: str | None = None,
        as_dict: bool = False,
    ) -> list[list[Any]] | list[dict[str, Any]]:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active

        rows = []
        for row in ws[range_spec]:
            rows.append([cell.value for cell in row])

        if as_dict and rows:
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            return [dict(zip(headers, row, strict=False)) for row in rows[1:]]

        logger.info(f"Read range {range_spec}: {len(rows)} rows")
        return rows

    @activity(name="Write Range", category="Excel")
    @tags("range", "write")
    def write_range(
        self,
        start_cell: str,
        data: list[list[Any]],
        sheet: str | None = None,
    ) -> None:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active

        start_col = ord(start_cell[0].upper()) - ord("A")
        start_row = (
            int("".join(filter(str.isdigit, start_cell)))
            if any(c.isdigit() for c in start_cell)
            else 1
        )

        for i, row_data in enumerate(data):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=start_col + j + 1)
                cell.value = value

        logger.info(f"Wrote range starting at {start_cell}: {len(data)} rows")

    @activity(name="Find Row", category="Excel")
    @tags("search", "row")
    @output("Row number if found, None otherwise")
    def find_row(
        self,
        column: int,
        value: Any,
        sheet: str | None = None,
    ) -> int | None:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active

        for row in ws.iter_rows(min_col=column, max_col=column):
            if row[0].value == value:
                return row[0].row
        return None

    @activity(name="Get Row Count", category="Excel")
    @tags("info", "row")
    @output("Number of rows")
    def get_row_count(self, sheet: str | None = None) -> int:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        return ws.max_row

    @activity(name="Get Column Count", category="Excel")
    @tags("info", "column")
    @output("Number of columns")
    def get_column_count(self, sheet: str | None = None) -> int:
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        return ws.max_column

    @activity(name="Insert Rows", category="Excel")
    @tags("row", "insert")
    def insert_rows(self, row: int, count: int = 1, sheet: str | None = None) -> None:
        """Insert blank rows above the given row number.

        :param row: Row number to insert above (1-based).
        :param count: Number of rows to insert.
        :param sheet: Sheet name (active sheet if None).
        """
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        ws.insert_rows(row, amount=count)
        logger.info(f"Inserted {count} row(s) above row {row}")

    @activity(name="Delete Rows", category="Excel")
    @tags("row", "delete")
    def delete_rows(self, row: int, count: int = 1, sheet: str | None = None) -> None:
        """Delete rows starting from the given row number.

        :param row: Starting row number (1-based).
        :param count: Number of rows to delete.
        :param sheet: Sheet name (active sheet if None).
        """
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        ws.delete_rows(row, amount=count)
        logger.info(f"Deleted {count} row(s) starting at row {row}")

    @activity(name="Insert Columns", category="Excel")
    @tags("column", "insert")
    def insert_columns(
        self, col: int, count: int = 1, sheet: str | None = None
    ) -> None:
        """Insert blank columns to the left of the given column number.

        :param col: Column number to insert before (1-based).
        :param count: Number of columns to insert.
        :param sheet: Sheet name (active sheet if None).
        """
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        ws.insert_cols(col, amount=count)
        logger.info(f"Inserted {count} column(s) before column {col}")

    @activity(name="Delete Columns", category="Excel")
    @tags("column", "delete")
    def delete_columns(
        self, col: int, count: int = 1, sheet: str | None = None
    ) -> None:
        """Delete columns starting from the given column number.

        :param col: Starting column number (1-based).
        :param count: Number of columns to delete.
        :param sheet: Sheet name (active sheet if None).
        """
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        ws.delete_cols(col, amount=count)
        logger.info(f"Deleted {count} column(s) starting at column {col}")

    @activity(name="Read Sheet To List", category="Excel")
    @tags("read", "sheet", "list")
    @output("List of row dicts keyed by header values")
    def read_sheet_to_list(
        self,
        sheet: str | None = None,
        header_row: int = 1,
        max_rows: int | None = None,
    ) -> list[dict[str, Any]]:
        """Read an entire sheet into a list of dicts using the header row as keys.

        :param sheet: Sheet name (active sheet if None).
        :param header_row: Row number containing column headers (1-based).
        :param max_rows: Maximum number of data rows to read (None for all rows).
        :returns: List of dicts, one per data row after the header.
        """
        if not self._workbook:
            raise ValueError("No workbook open")
        ws = self._workbook[sheet] if sheet else self._workbook.active
        max_row_arg = (header_row + max_rows) if max_rows is not None else None
        rows = list(ws.iter_rows(values_only=True, max_row=max_row_arg))
        if not rows:
            return []
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(rows[header_row - 1])
        ]
        result = [dict(zip(headers, row, strict=False)) for row in rows[header_row:]]
        logger.info(f"Read {len(result)} rows from sheet")
        return result

    @activity(name="Write List To Sheet", category="Excel")
    @tags("write", "sheet", "list")
    def write_list_to_sheet(
        self,
        data: list[dict[str, Any]],
        sheet: str | None = None,
        start_row: int = 1,
        write_headers: bool = True,
    ) -> None:
        """Write a list of dicts to the sheet, optionally writing a header row.

        :param data: List of dicts with identical keys.
        :param sheet: Sheet name (active sheet if None).
        :param start_row: Row to start writing (1-based).
        :param write_headers: Write column headers as the first row.
        """
        if not self._workbook:
            raise ValueError("No workbook open")
        if not data:
            return
        ws = self._workbook[sheet] if sheet else self._workbook.active
        headers = list(data[0].keys())
        row_num = start_row
        if write_headers:
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=row_num, column=col_num, value=header)
            row_num += 1
        for row_data in data:
            for col_num, key in enumerate(headers, start=1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(key))
            row_num += 1
        logger.info(f"Wrote {len(data)} rows to sheet")
